#!/usr/bin/env python3
"""Deterministic workbook stages for merging listed and clinical molecule pools."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
from collections import Counter, defaultdict
from copy import copy
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - environment failure
    raise SystemExit("This script requires openpyxl: python -m pip install openpyxl") from exc


META_SHEET = "__workflow_meta"
DEFAULT_SUPPLEMENT_AUDIT_SHEET = "__supplement_audit"
RELATION_HEADERS = [
    "判断编号",
    "临床来源行ID",
    "临床分子序号",
    "候选已上市来源行ID",
    "候选已上市分子序号",
    "临床登记号",
    "项目中国阶段",
    "品种中国阶段",
    "试验分期",
    "试验状态",
    "临床药品资产名称",
    "候选已上市药品资产名称",
    "候选已上市通用名",
    "临床剂型",
    "已上市剂型",
    "临床对应疾病",
    "已上市对应疾病",
    "临床相关适应症",
    "已上市相关适应症",
    "临床集团",
    "已上市集团",
    "名称匹配层级",
    "初始关系类型",
    "初始判断依据",
    "适应症核对结论",
    "Disease修正建议",
    "Disease修正值",
    "资产名称修正建议",
    "资产名称修正值",
    "修正后关系类型",
    "最终动作代码",
    "最终承接上市分子序号",
    "判断建议依据",
    "证据来源",
    "集中确认项",
    "判断状态",
    "主表同步状态",
    "阶段执行结果",
    "最终分子序号",
]
IMMUTABLE_RELATION_HEADERS = RELATION_HEADERS[:24]

ALLOWED_ACTIONS = {
    "KEEP",
    "DELETE_LISTED_COVERED",
    "EXCLUDE_NON_THERAPEUTIC",
    "EXCLUDE_NON_FIXED_REGIMEN",
    "EXCLUDE_PRE_SCOPE_MARKETED",
    "HOLD",
}

DOSAGE_TERMS = sorted(
    {
        "缓释胶囊",
        "控释胶囊",
        "分散片",
        "缓释片",
        "控释片",
        "咀嚼片",
        "肠溶片",
        "口服溶液剂",
        "冻干粉针剂",
        "注射液",
        "注射剂",
        "胶囊剂",
        "颗粒剂",
        "混悬剂",
        "吸入剂",
        "喷雾剂",
        "贴剂",
        "乳膏",
        "软膏",
        "片剂",
        "胶囊",
        "颗粒",
        "口服液",
        "注射用",
        "片",
    },
    key=len,
    reverse=True,
)


def emit(payload: Any) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


def load_json(path: str | Path) -> dict[str, Any]:
    p = Path(path)
    if not p.is_file():
        raise ValueError(f"JSON file was not found: {p}")
    with p.open("r", encoding="utf-8-sig") as handle:
        value = json.load(handle)
    if not isinstance(value, dict):
        raise ValueError(f"JSON root must be an object: {p}")
    return value


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_fingerprint(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def wide_schema(config: dict[str, Any]) -> list[dict[str, Any]]:
    schema = config.get("wide_schema")
    if not isinstance(schema, list) or not schema:
        raise ValueError("wide_schema must be a nonempty list.")
    return schema


def wide_fields(config: dict[str, Any]) -> dict[str, str]:
    fields = config.get("wide_fields")
    if not isinstance(fields, dict):
        raise ValueError("wide_fields must be an object.")
    return fields


def config_fingerprint(config: dict[str, Any]) -> str:
    contract = json.loads(json.dumps(config, ensure_ascii=False))
    for module in contract.get("downstream_modules", []):
        module.pop("status", None)
    return json_fingerprint(contract)


def run_fingerprint(config: dict[str, Any], listed_sha: str, clinical_sha: str) -> str:
    return json_fingerprint(
        {
            "config": config_fingerprint(config),
            "listed_sha256": listed_sha,
            "clinical_sha256": clinical_sha,
        }
    )


def require_new_output(output: str | Path, inputs: Iterable[str | Path] = ()) -> Path:
    out = Path(output).expanduser().resolve()
    if out.exists():
        raise ValueError(f"Output already exists and will not be overwritten: {out}")
    for item in inputs:
        if out == Path(item).expanduser().resolve():
            raise ValueError("Output path must differ from every input path.")
    if not out.parent.is_dir():
        raise ValueError(f"Output directory does not exist: {out.parent}")
    if out.suffix.lower() != ".xlsx":
        raise ValueError("Output workbook must use .xlsx.")
    return out


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def display_value(value: Any, missing: str) -> Any:
    return missing if is_blank(value) else value


def clean_text(value: Any, missing: str = "—") -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"", missing, "-", "--", "—", "——"}:
        return ""
    return text


def stable_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_carrier_sequences(value: Any) -> list[str]:
    text = stable_key(value)
    if not text:
        return []
    return [item for item in (part.strip() for part in re.split(r"[;；,/，\s]+", text)) if item]


def used_bounds(ws) -> tuple[int, int]:
    last_row = 1
    last_col = 1
    for row in ws.iter_rows():
        row_has_value = False
        for cell in row:
            if cell.value is not None:
                row_has_value = True
                last_col = max(last_col, cell.column)
        if row_has_value:
            last_row = max(last_row, row[0].row)
    return last_row, last_col


def header_map(ws) -> tuple[list[str], dict[str, int]]:
    _, last_col = used_bounds(ws)
    headers: list[str] = []
    mapping: dict[str, int] = {}
    for col in range(1, last_col + 1):
        name = clean_text(ws.cell(1, col).value, missing="")
        if not name:
            raise ValueError(f"Blank header at {ws.title}!{get_column_letter(col)}1")
        if name in mapping:
            raise ValueError(f"Duplicate header in {ws.title}: {name}")
        headers.append(name)
        mapping[name] = col
    return headers, mapping


def records_from_sheet(ws) -> tuple[list[str], dict[str, int], list[dict[str, Any]]]:
    headers, mapping = header_map(ws)
    last_row, _ = used_bounds(ws)
    records: list[dict[str, Any]] = []
    for row_number in range(2, last_row + 1):
        values = [ws.cell(row_number, col).value for col in range(1, len(headers) + 1)]
        if all(is_blank(value) for value in values):
            continue
        record = {header: values[index] for index, header in enumerate(headers)}
        record["__row_number__"] = row_number
        records.append(record)
    return headers, mapping, records


def records_with_cached_values(formula_ws, values_ws) -> tuple[list[str], dict[str, int], list[dict[str, Any]]]:
    """Use formula-sheet row presence and cached values for static mapped outputs."""
    headers, mapping = header_map(formula_ws)
    last_row, _ = used_bounds(formula_ws)
    records: list[dict[str, Any]] = []
    for row_number in range(2, last_row + 1):
        formula_values = [formula_ws.cell(row_number, col).value for col in range(1, len(headers) + 1)]
        if all(is_blank(value) for value in formula_values):
            continue
        cached_values = [values_ws.cell(row_number, col).value for col in range(1, len(headers) + 1)]
        record = {header: cached_values[index] for index, header in enumerate(headers)}
        record["__row_number__"] = row_number
        records.append(record)
    return headers, mapping, records


def sheet_fingerprint(ws, include_title: bool = True) -> str:
    last_row, last_col = used_bounds(ws)
    digest = hashlib.sha256()
    if include_title:
        digest.update(ws.title.encode("utf-8"))
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            if is_blank(cell.value):
                payload = f"{row}|{col}|BLANK\n"
            else:
                payload = f"{row}|{col}|{cell.data_type}|{type(cell.value).__name__}|{repr(cell.value)}\n"
            digest.update(payload.encode("utf-8", errors="replace"))
    for merged in sorted(str(item) for item in ws.merged_cells.ranges):
        digest.update(f"MERGED|{merged}\n".encode("utf-8"))
    return digest.hexdigest()


def formula_stats(ws) -> tuple[int, int]:
    last_row, last_col = used_bounds(ws)
    formulas = 0
    external = 0
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.startswith("="):
                formulas += 1
                if "[" in value:
                    external += 1
    return formulas, external


def validate_mapping_rule(rule: Any, label: str, allow_conditional: bool = False) -> None:
    if not isinstance(rule, dict):
        raise ValueError(f"Mapping rule must be an object: {label}")
    valid_modes = {"sequence", "constant", "column", "column_or_missing", "coalesce_columns", "missing"}
    if allow_conditional:
        valid_modes.add("conditional")
    mode = rule.get("mode")
    if mode not in valid_modes:
        raise ValueError(f"Invalid mapping mode for {label}: {mode}")
    if mode in {"column", "column_or_missing"} and not clean_text(rule.get("column"), missing=""):
        raise ValueError(f"Column mapping lacks a column name: {label}")
    if mode == "coalesce_columns":
        columns = rule.get("columns")
        if not isinstance(columns, list) or not columns or any(not clean_text(column, missing="") for column in columns):
            raise ValueError(f"Coalesced mapping needs nonempty columns: {label}")
    if mode == "conditional":
        cases = rule.get("cases")
        if not isinstance(cases, list) or not cases:
            raise ValueError(f"Conditional mapping needs cases: {label}")
        for index, case in enumerate(cases, start=1):
            if not isinstance(case, dict) or not isinstance(case.get("when"), dict):
                raise ValueError(f"Conditional case lacks a condition: {label} case {index}")
            validate_mapping_rule(case.get("rule"), f"{label} case {index}", allow_conditional=False)
        validate_mapping_rule(rule.get("default"), f"{label} default", allow_conditional=False)


def condition_references(condition: dict[str, Any]) -> set[str]:
    if "all" in condition:
        return set().union(*(condition_references(item) for item in condition["all"]))
    if "any" in condition:
        return set().union(*(condition_references(item) for item in condition["any"]))
    if "not" in condition:
        return condition_references(condition["not"])
    field = clean_text(condition.get("field"), missing="")
    return {field} if field else set()


def mapping_references(rule: dict[str, Any]) -> set[str]:
    mode = rule.get("mode")
    if mode in {"column", "column_or_missing"}:
        return {rule["column"]}
    if mode == "coalesce_columns":
        return set(rule["columns"])
    if mode == "conditional":
        result = mapping_references(rule["default"])
        for case in rule["cases"]:
            result |= condition_references(case["when"])
            result |= mapping_references(case["rule"])
        return result
    return set()


def validate_config_shape(config: dict[str, Any]) -> None:
    for section in ("listed", "clinical", "output", "wide_fields", "wide_schema", "clean_schema"):
        if section not in config:
            raise ValueError(f"Config is missing section: {section}")
    schema = wide_schema(config)
    names = [item.get("name") for item in schema]
    if any(not clean_text(name, missing="") for name in names) or len(names) != len(set(names)):
        raise ValueError("wide_schema field names must be nonblank and unique.")
    for item in schema:
        for source in ("listed", "clinical"):
            validate_mapping_rule(item.get(source), f"wide_schema {item.get('name')} / {source}")

    fields = wide_fields(config)
    required_logical_fields = {
        "sequence",
        "asset",
        "indication",
        "disease",
        "registration_no",
        "project_stage",
        "product_stage",
        "trial_phase",
        "trial_status",
        "common_name",
        "dosage_form",
        "group",
    }
    missing_logical_fields = sorted(required_logical_fields - set(fields))
    if missing_logical_fields:
        raise ValueError("wide_fields is missing logical fields: " + ", ".join(missing_logical_fields))
    wide_names = set(names)
    for logical, field in fields.items():
        if field and field not in wide_names:
            raise ValueError(f"wide_fields.{logical} points outside wide_schema: {field}")

    clean_items = config.get("clean_schema")
    auxiliary_items = config.get("clean_auxiliary_schema", [])
    if not isinstance(clean_items, list) or not clean_items:
        raise ValueError("clean_schema must be a nonempty list.")
    if not isinstance(auxiliary_items, list):
        raise ValueError("clean_auxiliary_schema must be a list.")
    clean_names = [item.get("name") for item in clean_items + auxiliary_items]
    if any(not clean_text(name, missing="") for name in clean_names) or len(clean_names) != len(set(clean_names)):
        raise ValueError("Clean field names must be nonblank and unique across core and auxiliary schemas.")
    for item in clean_items + auxiliary_items:
        rule = {key: value for key, value in item.items() if key != "name"}
        validate_mapping_rule(rule, f"clean field {item.get('name')}", allow_conditional=True)
        unknown_references = sorted(mapping_references(rule) - wide_names)
        if unknown_references:
            raise ValueError(f"Clean field {item.get('name')} references fields outside wide_schema: {unknown_references}")

    output = config["output"]
    for key in ("draft_sheet", "relation_sheet", "full_sheet", "clean_sheet", "listed_source_value", "clinical_source_value"):
        if not clean_text(output.get(key), missing=""):
            raise ValueError(f"output.{key} is required.")
    sheet_names = [output["draft_sheet"], output["relation_sheet"], output["full_sheet"], output["clean_sheet"], config["listed"]["raw_output_sheet"], config["clinical"]["raw_output_sheet"]]
    if len(sheet_names) != len(set(sheet_names)):
        raise ValueError("Configured output sheet names must be unique.")

    supplement = config.get("supplement", {})
    if supplement.get("enabled"):
        for key in ("template_sheet", "id_field", "status_field", "open_status", "closed_statuses", "trigger", "target_fields"):
            if key not in supplement:
                raise ValueError(f"supplement.{key} is required when supplementation is enabled.")
        if supplement["id_field"] not in wide_names:
            raise ValueError("supplement.id_field must exist in wide_schema.")
        for field in supplement.get("context_fields", []) + supplement.get("target_fields", []):
            if field not in wide_names:
                raise ValueError(f"Supplement field points outside wide_schema: {field}")
        unknown_trigger_fields = sorted(condition_references(supplement["trigger"]) - wide_names)
        if unknown_trigger_fields:
            raise ValueError(f"Supplement trigger references fields outside wide_schema: {unknown_trigger_fields}")
        if not isinstance(supplement["closed_statuses"], list) or not supplement["closed_statuses"]:
            raise ValueError("supplement.closed_statuses must be a nonempty list.")

    modules = config.get("downstream_modules", [])
    if not isinstance(modules, list):
        raise ValueError("downstream_modules must be a list.")
    allowed_module_statuses = {"pending", "completed", "not_applicable", "approved_skip"}
    for module in modules:
        if not isinstance(module, dict) or not clean_text(module.get("name"), missing=""):
            raise ValueError("Every downstream module needs a name.")
        if module.get("status", "pending") not in allowed_module_statuses:
            raise ValueError(f"Invalid downstream module status: {module.get('name')}")
        unknown_affected = sorted(set(module.get("affected_by", [])) - wide_names)
        if unknown_affected:
            raise ValueError(f"Module {module.get('name')} affected_by fields are outside wide_schema: {unknown_affected}")


def source_required_fields(config: dict[str, Any], source: str) -> list[str]:
    section = config[source]
    logical = ["asset", "indication", "disease", "dosage_form"]
    if source == "clinical":
        logical.extend(["project_stage", "product_stage", "trial_phase", "trial_status"])
    required = [section.get(item, "") for item in logical]
    source_id = section.get("source_id", "")
    if source_id:
        required.append(source_id)
    for item in wide_schema(config):
        rule = item[source]
        if rule["mode"] == "column":
            required.append(rule["column"])
    return sorted({item for item in required if item})


def inspect_one_source(config: dict[str, Any], source: str) -> tuple[dict[str, Any], list[str]]:
    section = config[source]
    path = Path(section["file"])
    blockers: list[str] = []
    if not path.is_file():
        return {"file": str(path), "sheet": section.get("sheet"), "readable": False}, [f"Missing {source} file: {path}"]
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        blockers.append(f"Unsupported {source} workbook extension: {path.suffix}")
    wb = load_workbook(path, data_only=False, read_only=False)
    sheet_name = section["sheet"]
    if sheet_name not in wb.sheetnames:
        return {"file": str(path), "sheet": sheet_name, "readable": False, "sheets": wb.sheetnames}, blockers + [f"Missing {source} sheet: {sheet_name}"]
    ws = wb[sheet_name]
    try:
        headers, _, records = records_from_sheet(ws)
    except ValueError as exc:
        return {"file": str(path), "sheet": sheet_name, "readable": False}, blockers + [str(exc)]
    missing_required = [field for field in source_required_fields(config, source) if field not in headers]
    blockers.extend(f"{source} missing required field: {field}" for field in missing_required)
    for item in wide_schema(config):
        rule = item[source]
        if rule["mode"] == "coalesce_columns" and not any(column in headers for column in rule["columns"]):
            blockers.append(
                f"{source} lacks every fallback column for {item['name']}: {', '.join(rule['columns'])}"
            )

    source_id = section.get("source_id", "")
    blank_ids = 0
    duplicate_ids: list[str] = []
    generated_ids = not source_id or source_id not in headers
    if not generated_ids:
        values = [stable_key(record.get(source_id)) for record in records]
        blank_ids = sum(not value for value in values)
        counts = Counter(value for value in values if value)
        duplicate_ids = [value for value, count in counts.items() if count > 1]
        if blank_ids:
            blockers.append(f"{source} has {blank_ids} blank source IDs")
        if duplicate_ids:
            blockers.append(f"{source} has {len(duplicate_ids)} duplicate source IDs")

    blank_by_field: dict[str, int] = {}
    for logical in ("asset", "indication", "disease", "dosage_form"):
        field = section.get(logical, "")
        if field in headers:
            blank_by_field[field] = sum(is_blank(record.get(field)) for record in records)
    disease_field = section.get("disease", "")
    multi_disease_rows: list[int] = []
    if disease_field in headers:
        for record in records:
            value = clean_text(record.get(disease_field))
            if re.search(r"[;；\r\n]", value):
                multi_disease_rows.append(record["__row_number__"])
        if blank_by_field.get(disease_field, 0):
            blockers.append(f"{source} has blank Disease values: {blank_by_field[disease_field]}")
        if multi_disease_rows:
            blockers.append(f"{source} has unresolved multivalue Disease rows: {len(multi_disease_rows)}")

    duplicate_signatures: dict[str, list[int]] = defaultdict(list)
    asset = section.get("asset", "")
    form = section.get("dosage_form", "")
    if all(field in headers for field in (asset, form, disease_field)):
        for record in records:
            signature = " | ".join(clean_text(record.get(field)).casefold() for field in (asset, form, disease_field))
            duplicate_signatures[signature].append(record["__row_number__"])
    duplicate_signatures = {key: rows for key, rows in duplicate_signatures.items() if key.strip(" |") and len(rows) > 1}
    if duplicate_signatures and not section.get("within_source_duplicates_approved", False):
        blockers.append(
            f"{source} has {len(duplicate_signatures)} within-source asset/form/Disease duplicate signatures that require upstream resolution or explicit approval"
        )

    status_counts: dict[str, dict[str, int]] = {}
    if source == "clinical":
        for logical in ("project_stage", "product_stage", "trial_phase", "trial_status"):
            field = section.get(logical, "")
            if field in headers:
                status_counts[field] = dict(Counter(clean_text(record.get(field)) for record in records))
    formulas, external = formula_stats(ws)
    cached_missing = 0
    cached_missing_required = 0
    if formulas:
        values_wb = load_workbook(path, data_only=True, read_only=False)
        values_ws = values_wb[sheet_name]
        required_columns = {mapping[field] for field in source_required_fields(config, source) if field in mapping}
        last_formula_row, last_formula_col = used_bounds(ws)
        for row in range(2, last_formula_row + 1):
            for col in range(1, last_formula_col + 1):
                value = ws.cell(row, col).value
                if isinstance(value, str) and value.startswith("=") and values_ws.cell(row, col).value is None:
                    cached_missing += 1
                    if col in required_columns:
                        cached_missing_required += 1
        values_wb.close()
    if cached_missing_required:
        blockers.append(f"{source} has {cached_missing_required} required mapped formula cells without cached values")
    last_row, last_col = used_bounds(ws)
    result = {
        "file": str(path.resolve()),
        "sheet": sheet_name,
        "readable": True,
        "data_rows": len(records),
        "used_rows_including_header": last_row,
        "columns": last_col,
        "headers": headers,
        "missing_required_fields": missing_required,
        "source_id_field": source_id or None,
        "generated_source_ids": generated_ids,
        "blank_source_ids": blank_ids,
        "duplicate_source_ids": duplicate_ids,
        "blank_key_fields": blank_by_field,
        "multivalue_disease_rows": multi_disease_rows,
        "within_source_duplicate_signature_count": len(duplicate_signatures),
        "within_source_duplicate_signatures": duplicate_signatures,
        "status_counts": status_counts,
        "formula_cells": formulas,
        "potential_external_formula_cells": external,
        "formula_cells_without_cached_values": cached_missing,
        "required_formula_cells_without_cached_values": cached_missing_required,
        "file_sha256": file_sha256(path),
        "fingerprint": sheet_fingerprint(ws),
    }
    return result, blockers


def inspect_config(config: dict[str, Any]) -> dict[str, Any]:
    validate_config_shape(config)
    listed, listed_blockers = inspect_one_source(config, "listed")
    clinical, clinical_blockers = inspect_one_source(config, "clinical")
    run_schema = [item["name"] for item in wide_schema(config)]
    return {
        "state": "PREFLIGHT_PASSED" if not (listed_blockers + clinical_blockers) else "PREFLIGHT_FAILED",
        "ta_name": config.get("ta_name"),
        "schema_approved": bool(config.get("schema_approved")),
        "wide_schema_field_count": len(run_schema),
        "wide_schema": run_schema,
        "clean_schema_field_count": len(config["clean_schema"]),
        "clean_schema": [item["name"] for item in config["clean_schema"]],
        "clean_auxiliary_field_count": len(config.get("clean_auxiliary_schema", [])),
        "config_fingerprint": config_fingerprint(config),
        "listed": listed,
        "clinical": clinical,
        "blockers": listed_blockers + clinical_blockers,
    }


def copy_sheet(source_ws, target_wb, target_name: str):
    target = target_wb.create_sheet(target_name)
    last_row, last_col = used_bounds(source_ws)
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            source_cell = source_ws.cell(row, col)
            target_cell = target.cell(row, col, source_cell.value)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.protection = copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
    for key, dimension in source_ws.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
        target.column_dimensions[key].outlineLevel = dimension.outlineLevel
    for key, dimension in source_ws.row_dimensions.items():
        target.row_dimensions[key].height = dimension.height
        target.row_dimensions[key].hidden = dimension.hidden
        target.row_dimensions[key].outlineLevel = dimension.outlineLevel
    for merged in source_ws.merged_cells.ranges:
        target.merge_cells(str(merged))
    target.freeze_panes = source_ws.freeze_panes
    target.auto_filter.ref = source_ws.auto_filter.ref
    target.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target.sheet_properties = copy(source_ws.sheet_properties)
    target.sheet_format = copy(source_ws.sheet_format)
    target.page_margins = copy(source_ws.page_margins)
    target.page_setup = copy(source_ws.page_setup)
    target.print_options = copy(source_ws.print_options)
    target.data_validations = copy(source_ws.data_validations)
    target.conditional_formatting = copy(source_ws.conditional_formatting)
    return target


def style_table(ws, header_fill: str = "1F4E78") -> None:
    if ws.max_column < 1:
        return
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        width = max(10, min(32, max(len(str(ws.cell(row, col).value or "")) for row in range(1, min(ws.max_row, 50) + 1)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def get_source_id(record: dict[str, Any], section: dict[str, Any], raw_sheet: str) -> str:
    source_id = section.get("source_id", "")
    if source_id and not is_blank(record.get(source_id)):
        return stable_key(record[source_id])
    return f"{raw_sheet}!R{record['__row_number__']}"


def mapped_value(rule: dict[str, Any], record: dict[str, Any], sequence: int, missing: str) -> Any:
    mode = rule["mode"]
    if mode == "sequence":
        return sequence
    if mode == "constant":
        return display_value(rule.get("value"), missing)
    if mode == "missing":
        return missing
    if mode == "coalesce_columns":
        for column in rule["columns"]:
            if column in record and not is_blank(record.get(column)):
                return display_value(record.get(column), missing)
        return missing
    column = rule["column"]
    if column not in record:
        if mode == "column_or_missing":
            return missing
        raise ValueError(f"Required mapped column is missing: {column}")
    return display_value(record.get(column), missing)


def set_meta_value(ws, key: str, value: Any) -> None:
    last_row = max(used_bounds(ws)[0], 1)
    for row in range(1, last_row + 1):
        if clean_text(ws.cell(row, 6).value, missing="") == key:
            ws.cell(row, 7).value = value
            return
    row = last_row + 1
    ws.cell(row, 6).value = key
    ws.cell(row, 7).value = value


def meta_values(ws) -> dict[str, str]:
    result: dict[str, str] = {}
    last_row = max(used_bounds(ws)[0], 1)
    for row in range(1, last_row + 1):
        key = clean_text(ws.cell(row, 6).value, missing="")
        if key:
            result[key] = clean_text(ws.cell(row, 7).value, missing="")
    return result


def require_matching_run(config: dict[str, Any], meta: dict[str, str]) -> None:
    expected_config = config_fingerprint(config)
    if meta.get("config_fingerprint") != expected_config:
        raise ValueError("Run config fingerprint differs from the workbook manifest.")
    for source in ("listed", "clinical"):
        path = Path(config[source]["file"])
        if not path.is_file():
            raise ValueError(f"Configured {source} input is no longer available: {path}")
        expected_sha = meta.get(f"{source}_input_sha256")
        if expected_sha and file_sha256(path) != expected_sha:
            raise ValueError(f"Configured {source} input SHA-256 differs from the workbook manifest.")


def command_inspect(args) -> None:
    config = load_json(args.config)
    emit(inspect_config(config))


def command_build_draft(args) -> None:
    config = load_json(args.config)
    preflight = inspect_config(config)
    if preflight["blockers"]:
        raise ValueError("Preflight failed: " + " | ".join(preflight["blockers"]))
    if not args.confirmed_schema or not config.get("schema_approved"):
        raise ValueError("Draft generation requires explicit Schema approval and schema_approved=true.")
    listed_path = Path(config["listed"]["file"])
    clinical_path = Path(config["clinical"]["file"])
    output = require_new_output(args.output, [listed_path, clinical_path])
    listed_wb = load_workbook(listed_path, data_only=False)
    clinical_wb = load_workbook(clinical_path, data_only=False)
    listed_values_wb = load_workbook(listed_path, data_only=True)
    clinical_values_wb = load_workbook(clinical_path, data_only=True)
    listed_ws = listed_wb[config["listed"]["sheet"]]
    clinical_ws = clinical_wb[config["clinical"]["sheet"]]
    listed_values_ws = listed_values_wb[config["listed"]["sheet"]]
    clinical_values_ws = clinical_values_wb[config["clinical"]["sheet"]]
    _, _, listed_records = records_with_cached_values(listed_ws, listed_values_ws)
    _, _, clinical_records = records_with_cached_values(clinical_ws, clinical_values_ws)

    wb = Workbook()
    wb.remove(wb.active)
    draft_name = config["output"]["draft_sheet"]
    draft = wb.create_sheet(draft_name)
    run_schema = wide_schema(config)
    headers = [item["name"] for item in run_schema]
    draft.append(headers)
    meta = wb.create_sheet(META_SHEET)
    meta.append(["数据来源", "待判断分子序号", "来源行ID", "原始Sheet行号"])
    missing = config.get("missing_display", "—")
    listed_source = config["output"]["listed_source_value"]
    clinical_source = config["output"]["clinical_source_value"]
    sequence = 1

    for source, records, section, source_value in (
        ("listed", listed_records, config["listed"], listed_source),
        ("clinical", clinical_records, config["clinical"], clinical_source),
    ):
        raw_name = section["raw_output_sheet"]
        for record in records:
            row = [mapped_value(item[source], record, sequence, missing) for item in run_schema]
            draft.append(row)
            meta.append([source_value, sequence, get_source_id(record, section, raw_name), record["__row_number__"]])
            sequence += 1

    listed_original_fp = sheet_fingerprint(listed_ws, include_title=False)
    clinical_original_fp = sheet_fingerprint(clinical_ws, include_title=False)
    listed_raw = copy_sheet(listed_ws, wb, config["listed"]["raw_output_sheet"])
    clinical_raw = copy_sheet(clinical_ws, wb, config["clinical"]["raw_output_sheet"])
    listed_sha = file_sha256(listed_path)
    clinical_sha = file_sha256(clinical_path)
    set_meta_value(meta, "state", "DRAFT_BUILT")
    set_meta_value(meta, "skill_version", str(config.get("skill_version", "2.0")))
    set_meta_value(meta, "config_fingerprint", config_fingerprint(config))
    set_meta_value(meta, "run_fingerprint", run_fingerprint(config, listed_sha, clinical_sha))
    set_meta_value(meta, "listed_input_sha256", listed_sha)
    set_meta_value(meta, "clinical_input_sha256", clinical_sha)
    set_meta_value(meta, "listed_raw_fingerprint", listed_original_fp)
    set_meta_value(meta, "clinical_raw_fingerprint", clinical_original_fp)
    set_meta_value(meta, "wide_schema", json.dumps(headers, ensure_ascii=False))
    set_meta_value(meta, "clean_schema", json.dumps([item["name"] for item in config["clean_schema"]], ensure_ascii=False))
    set_meta_value(meta, "draft_fingerprint", sheet_fingerprint(draft))
    set_meta_value(meta, "clean_state", "NOT_GENERATED")
    meta.sheet_state = "hidden"
    style_table(draft)
    wb.save(output)

    verify = load_workbook(output, data_only=False)
    verify_draft = verify[draft_name]
    _, _, verify_rows = records_from_sheet(verify_draft)
    listed_fp = sheet_fingerprint(verify[config["listed"]["raw_output_sheet"]], include_title=False)
    clinical_fp = sheet_fingerprint(verify[config["clinical"]["raw_output_sheet"]], include_title=False)
    expected = len(listed_records) + len(clinical_records)
    if len(verify_rows) != expected:
        output.unlink(missing_ok=True)
        raise ValueError("Draft row-count verification failed.")
    if listed_fp != listed_original_fp or clinical_fp != clinical_original_fp:
        output.unlink(missing_ok=True)
        raise ValueError("Raw-sheet fingerprint verification failed after draft generation.")
    emit(
        {
            "state": "DRAFT_BUILT",
            "output": str(output),
            "listed_rows": len(listed_records),
            "clinical_rows": len(clinical_records),
            "draft_rows": len(verify_rows),
            "wide_schema_fields": len(headers),
            "clean_schema_fields": len(config["clean_schema"]),
            "config_fingerprint": config_fingerprint(config),
            "run_fingerprint": run_fingerprint(config, listed_sha, clinical_sha),
            "raw_sheet_cell_formula_fingerprints_preserved": True,
            "hidden_meta_sheet": META_SHEET,
        }
    )
    listed_values_wb.close()
    clinical_values_wb.close()


def normalize_name(value: Any, missing: str) -> str:
    text = clean_text(value, missing).casefold().replace("＋", "+")
    return re.sub(r"[\s·•,，;；:：()（）\[\]【】{}\-_/\\+]", "", text)


def strip_dosage_form(value: Any, dosage_form: Any, missing: str) -> str:
    base = normalize_name(value, missing)
    form = normalize_name(dosage_form, missing)
    if form and base.endswith(form) and len(base) > len(form):
        base = base[: -len(form)]
    for term in DOSAGE_TERMS:
        normalized_term = normalize_name(term, missing)
        if normalized_term and base.endswith(normalized_term) and len(base) > len(normalized_term):
            base = base[: -len(normalized_term)]
            break
    return base


def aliases(asset: Any, common: Any, dosage_form: Any, missing: str) -> set[str]:
    result = {
        normalize_name(asset, missing),
        strip_dosage_form(asset, dosage_form, missing),
        normalize_name(common, missing),
        strip_dosage_form(common, dosage_form, missing),
    }
    return {item for item in result if item}


def components(asset: Any, dosage_form: Any, missing: str) -> set[str]:
    raw = clean_text(asset, missing)
    parts = re.split(r"[+＋;；、,，\r\n/&]", raw)
    result = {strip_dosage_form(part, dosage_form, missing) for part in parts}
    return {item for item in result if item}


def normalized_equal(left: Any, right: Any, missing: str) -> bool:
    lvalue = normalize_name(left, missing)
    rvalue = normalize_name(right, missing)
    return bool(lvalue and rvalue and lvalue == rvalue)


def read_meta(ws) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    mapping: dict[str, dict[str, str]] = {}
    last_row, _ = used_bounds(ws)
    for row in range(2, last_row + 1):
        source = clean_text(ws.cell(row, 1).value)
        sequence = stable_key(ws.cell(row, 2).value)
        if source and sequence:
            mapping[sequence] = {
                "source": source,
                "source_id": stable_key(ws.cell(row, 3).value),
                "raw_row": stable_key(ws.cell(row, 4).value),
            }
    state: dict[str, str] = {}
    for row in range(1, max(last_row, 10) + 1):
        key = clean_text(ws.cell(row, 6).value)
        if key:
            state[key] = clean_text(ws.cell(row, 7).value, missing="")
    return mapping, state


def record_source(
    record: dict[str, Any],
    fields: dict[str, str],
    meta_map: dict[str, dict[str, str]],
) -> str:
    """Return source provenance without requiring a visible output column."""
    source_field = fields.get("source", "")
    if source_field and source_field in record:
        source = clean_text(record.get(source_field), missing="")
        if source:
            return source
    sequence = stable_key(record.get(fields["sequence"]))
    return clean_text(meta_map.get(sequence, {}).get("source"), missing="")


def get_record(record: dict[str, Any], field_name: str) -> Any:
    return record.get(field_name, "")


def direct_candidates(clinical: dict[str, Any], listed: list[dict[str, Any]], fields: dict[str, str], missing: str):
    c_alias = aliases(get_record(clinical, fields["asset"]), "", get_record(clinical, fields["dosage_form"]), missing)
    result = []
    for listed_record in listed:
        l_alias = aliases(
            get_record(listed_record, fields["asset"]),
            get_record(listed_record, fields["common_name"]),
            get_record(listed_record, fields["dosage_form"]),
            missing,
        )
        if c_alias & l_alias:
            result.append((listed_record, "direct", 1.0))
    return result


def partial_candidates(clinical: dict[str, Any], listed: list[dict[str, Any]], fields: dict[str, str], missing: str):
    c_parts = components(get_record(clinical, fields["asset"]), get_record(clinical, fields["dosage_form"]), missing)
    result = []
    for listed_record in listed:
        l_alias = aliases(
            get_record(listed_record, fields["asset"]),
            get_record(listed_record, fields["common_name"]),
            get_record(listed_record, fields["dosage_form"]),
            missing,
        )
        if c_parts & l_alias:
            result.append((listed_record, "partial", 0.8))
    return result


def approximate_candidates(
    clinical: dict[str, Any],
    listed: list[dict[str, Any]],
    fields: dict[str, str],
    missing: str,
    threshold: float,
    limit: int,
):
    clinical_name = strip_dosage_form(get_record(clinical, fields["asset"]), get_record(clinical, fields["dosage_form"]), missing)
    if len(clinical_name) < 4:
        return []
    scored = []
    for listed_record in listed:
        l_alias = aliases(
            get_record(listed_record, fields["asset"]),
            get_record(listed_record, fields["common_name"]),
            get_record(listed_record, fields["dosage_form"]),
            missing,
        )
        score = max((SequenceMatcher(None, clinical_name, alias).ratio() for alias in l_alias if len(alias) >= 4), default=0.0)
        if score >= threshold:
            scored.append((listed_record, "approximate", score))
    return sorted(scored, key=lambda item: item[2], reverse=True)[:limit]


def initial_relation(clinical: dict[str, Any], listed: dict[str, Any] | None, match_type: str, fields: dict[str, str], missing: str) -> tuple[str, str, str]:
    if listed is None:
        stages = " ".join(
            clean_text(get_record(clinical, fields[item]), missing)
            for item in ("project_stage", "product_stage")
            if item in fields
        )
        relation = "已上市但当前上市池无直接候选" if "已上市" in stages else "临床独有（无直接候选）"
        return "无直接名称候选", relation, "No listed product/generic candidate was generated; review scope and marketed-gap evidence if relevant."
    same_form = normalized_equal(get_record(clinical, fields["dosage_form"]), get_record(listed, fields["dosage_form"]), missing)
    same_disease = clean_text(get_record(clinical, fields["disease"]), missing) == clean_text(get_record(listed, fields["disease"]), missing)
    if match_type == "partial":
        return "部分成分与上市资产一致", "部分成分重合，非同一资产", "At least one normalized clinical component matches a listed product or generic name; inspect the complete component set."
    if match_type == "approximate":
        return "名称近似但关键表达可能不同", "近似名称/盐型或化学形式差异", "Name similarity generated this review candidate; do not treat it as identity evidence."
    if same_form and same_disease:
        relation = "高度疑似重复"
    elif same_form:
        relation = "同分子同剂型，不同疾病"
    elif same_disease:
        relation = "同分子不同剂型，同疾病"
    else:
        relation = "同分子，剂型与疾病均不同"
    return "名称与上市药品/通用名直接一致", relation, "Normalized clinical asset matched a listed product or generic name; dosage form and Disease were compared separately."


def command_build_candidates(args) -> None:
    config = load_json(args.config)
    validate_config_shape(config)
    if not config.get("schema_approved"):
        raise ValueError("Candidate generation requires an approved Run Schema.")
    source = Path(args.workbook)
    if not source.is_file():
        raise ValueError(f"Draft workbook was not found: {source}")
    output = require_new_output(args.output, [source])
    shutil.copy2(source, output)
    wb = load_workbook(output, data_only=False)
    draft_name = config["output"]["draft_sheet"]
    relation_name = config["output"]["relation_sheet"]
    if draft_name not in wb.sheetnames or META_SHEET not in wb.sheetnames:
        output.unlink(missing_ok=True)
        raise ValueError("Workbook lacks the draft or workflow metadata sheet.")
    if relation_name in wb.sheetnames:
        output.unlink(missing_ok=True)
        raise ValueError(f"Relation sheet already exists: {relation_name}")
    draft = wb[draft_name]
    headers, _, records = records_from_sheet(draft)
    expected_headers = [item["name"] for item in wide_schema(config)]
    if headers != expected_headers:
        output.unlink(missing_ok=True)
        raise ValueError("Draft headers do not equal the approved Run Schema.")
    meta_map, meta_state = read_meta(wb[META_SHEET])
    require_matching_run(config, meta_state)
    if meta_state.get("draft_fingerprint") != sheet_fingerprint(draft):
        output.unlink(missing_ok=True)
        raise ValueError("Draft fingerprint differs from the workbook manifest.")
    fields = wide_fields(config)
    sequence_field = fields["sequence"]
    listed_source = config["output"]["listed_source_value"]
    clinical_source = config["output"]["clinical_source_value"]
    listed = [record for record in records if record_source(record, fields, meta_map) == listed_source]
    clinical = [record for record in records if record_source(record, fields, meta_map) == clinical_source]
    missing = config.get("missing_display", "—")
    threshold = float(config.get("candidate_similarity_threshold", 0.78))
    approx_limit = int(config.get("approximate_candidate_limit", 5))
    relation = wb.create_sheet(relation_name, 1)
    relation.append(RELATION_HEADERS)
    relation_index = {name: index for index, name in enumerate(RELATION_HEADERS)}
    relation_rows = 0
    one_to_many = 0
    relation_counts: Counter[str] = Counter()

    for clinical_record in clinical:
        candidates = direct_candidates(clinical_record, listed, fields, missing)
        if not candidates:
            candidates = partial_candidates(clinical_record, listed, fields, missing)
        if not candidates:
            candidates = approximate_candidates(clinical_record, listed, fields, missing, threshold, approx_limit)
        if not candidates:
            candidates = [(None, "none", 0.0)]
        if len(candidates) > 1:
            one_to_many += 1
        clinical_sequence = stable_key(clinical_record[sequence_field])
        clinical_meta = meta_map.get(clinical_sequence, {})
        for listed_record, match_type, score in candidates:
            relation_rows += 1
            match_level, relation_type, basis = initial_relation(clinical_record, listed_record, match_type, fields, missing)
            if match_type == "approximate":
                basis += f" Similarity score={score:.3f}."
            relation_counts[relation_type] += 1
            row = [""] * len(RELATION_HEADERS)

            def put(name: str, value: Any) -> None:
                row[relation_index[name]] = "" if value is None else value

            put("判断编号", relation_rows)
            put("临床来源行ID", clinical_meta.get("source_id", f"临床序号:{clinical_sequence}"))
            put("临床分子序号", clinical_record[sequence_field])
            put("临床登记号", get_record(clinical_record, fields["registration_no"]))
            put("项目中国阶段", get_record(clinical_record, fields["project_stage"]))
            put("品种中国阶段", get_record(clinical_record, fields["product_stage"]))
            put("试验分期", get_record(clinical_record, fields["trial_phase"]))
            put("试验状态", get_record(clinical_record, fields["trial_status"]))
            put("临床药品资产名称", get_record(clinical_record, fields["asset"]))
            put("临床剂型", get_record(clinical_record, fields["dosage_form"]))
            put("临床对应疾病", get_record(clinical_record, fields["disease"]))
            put("临床相关适应症", get_record(clinical_record, fields["indication"]))
            put("临床集团", get_record(clinical_record, fields["group"]))
            if listed_record is not None:
                listed_sequence = stable_key(listed_record[sequence_field])
                listed_meta = meta_map.get(listed_sequence, {})
                put("候选已上市来源行ID", listed_meta.get("source_id", f"上市序号:{listed_sequence}"))
                put("候选已上市分子序号", listed_record[sequence_field])
                put("候选已上市药品资产名称", get_record(listed_record, fields["asset"]))
                put("候选已上市通用名", get_record(listed_record, fields["common_name"]))
                put("已上市剂型", get_record(listed_record, fields["dosage_form"]))
                put("已上市对应疾病", get_record(listed_record, fields["disease"]))
                put("已上市相关适应症", get_record(listed_record, fields["indication"]))
                put("已上市集团", get_record(listed_record, fields["group"]))
            put("名称匹配层级", match_level)
            put("初始关系类型", relation_type)
            put("初始判断依据", basis)
            put("判断状态", "待判断")
            relation.append(row)

    style_table(relation, header_fill="7030A0")
    candidate_fp = candidate_fingerprint(relation)
    set_meta_value(wb[META_SHEET], "state", "RELATIONS_BUILT")
    set_meta_value(wb[META_SHEET], "candidate_fingerprint", candidate_fp)
    wb.save(output)
    emit(
        {
            "state": "RELATIONS_BUILT",
            "output": str(output),
            "listed_rows": len(listed),
            "unique_clinical_rows": len(clinical),
            "candidate_rows": relation_rows,
            "one_to_many_clinical_rows": one_to_many,
            "initial_relation_types": dict(relation_counts),
            "candidate_fingerprint": candidate_fp,
            "automatic_deletions": 0,
        }
    )


def relation_header_map(ws) -> dict[str, int]:
    headers, mapping = header_map(ws)
    missing = [header for header in RELATION_HEADERS if header not in mapping]
    if missing:
        raise ValueError("Relation sheet lacks required columns: " + ", ".join(missing))
    return mapping


def candidate_fingerprint(ws) -> str:
    mapping = relation_header_map(ws)
    last_row, _ = used_bounds(ws)
    digest = hashlib.sha256()
    for header in IMMUTABLE_RELATION_HEADERS:
        digest.update((header + "\n").encode("utf-8"))
    for row in range(2, last_row + 1):
        for header in IMMUTABLE_RELATION_HEADERS:
            value = ws.cell(row, mapping[header]).value
            if is_blank(value):
                normalized = ""
            elif isinstance(value, float) and value.is_integer():
                normalized = str(int(value))
            else:
                normalized = str(value)
            digest.update(f"{row}|{header}|{normalized}\n".encode("utf-8", errors="replace"))
    return digest.hexdigest()


def command_import_decisions(args) -> None:
    config = load_json(args.config)
    validate_config_shape(config)
    source = Path(args.workbook)
    if not source.is_file():
        raise ValueError(f"Candidate workbook was not found: {source}")
    output = require_new_output(args.output, [source])
    payload = load_json(args.decisions)
    decisions = payload.get("decisions")
    if not isinstance(decisions, list):
        raise ValueError("Decision JSON must contain a decisions list.")
    shutil.copy2(source, output)
    wb = load_workbook(output, data_only=False)
    relation_name = config["output"]["relation_sheet"]
    if relation_name not in wb.sheetnames or META_SHEET not in wb.sheetnames:
        output.unlink(missing_ok=True)
        raise ValueError("Relation or workflow metadata sheet was not found.")
    ws = wb[relation_name]
    meta = meta_values(wb[META_SHEET])
    require_matching_run(config, meta)
    current_candidate_fp = candidate_fingerprint(ws)
    if meta.get("candidate_fingerprint") != current_candidate_fp:
        output.unlink(missing_ok=True)
        raise ValueError("Candidate fingerprint differs from the workbook manifest.")
    if clean_text(payload.get("candidate_fingerprint"), missing="") != current_candidate_fp:
        output.unlink(missing_ok=True)
        raise ValueError("Decision JSON candidate_fingerprint does not match the workbook.")
    mapping = relation_header_map(ws)
    last_row, _ = used_bounds(ws)
    by_source: dict[str, list[int]] = defaultdict(list)
    by_sequence: dict[str, list[int]] = defaultdict(list)
    for row in range(2, last_row + 1):
        by_source[stable_key(ws.cell(row, mapping["临床来源行ID"]).value)].append(row)
        by_sequence[stable_key(ws.cell(row, mapping["临床分子序号"]).value)].append(row)
    assigned: set[str] = set()
    field_map = {
        "indication_check": "适应症核对结论",
        "disease_advice": "Disease修正建议",
        "disease_correction": "Disease修正值",
        "asset_name_advice": "资产名称修正建议",
        "asset_name_correction": "资产名称修正值",
        "revised_relation_type": "修正后关系类型",
        "final_action": "最终动作代码",
        "carrier_listed_sequence": "最终承接上市分子序号",
        "rationale": "判断建议依据",
        "evidence_source": "证据来源",
        "confirmation_group": "集中确认项",
        "decision_status": "判断状态",
    }
    for decision in decisions:
        if not isinstance(decision, dict):
            raise ValueError("Each decision must be an object.")
        source_id = stable_key(decision.get("clinical_source_id"))
        sequence = stable_key(decision.get("clinical_sequence"))
        source_rows = by_source.get(source_id, []) if source_id else []
        sequence_rows = by_sequence.get(sequence, []) if sequence else []
        if source_id and sequence and set(source_rows) != set(sequence_rows):
            raise ValueError(f"Decision source ID and sequence identify different groups: {source_id} / {sequence}")
        rows = source_rows or sequence_rows
        if not rows:
            raise ValueError(f"Decision did not match a clinical group: {source_id or sequence}")
        group_key = stable_key(ws.cell(rows[0], mapping["临床分子序号"]).value)
        if group_key in assigned:
            raise ValueError(f"Duplicate decision for clinical sequence: {group_key}")
        assigned.add(group_key)
        action = clean_text(decision.get("final_action"), missing="")
        if action and action not in ALLOWED_ACTIONS:
            raise ValueError(f"Unsupported action: {action}")
        for row in rows:
            for source_field, relation_field in field_map.items():
                if source_field in decision:
                    ws.cell(row, mapping[relation_field]).value = decision[source_field]
            has_correction = bool(clean_text(decision.get("disease_correction"), "") or clean_text(decision.get("asset_name_correction"), ""))
            ws.cell(row, mapping["主表同步状态"]).value = "待同步" if has_correction else "无需修正"
    unique_sequences = set(by_sequence)
    pending = unique_sequences - assigned
    final_state = "DECISIONS_CLOSED" if not pending else "DECISIONS_PENDING"
    set_meta_value(wb[META_SHEET], "state", final_state)
    set_meta_value(wb[META_SHEET], "decision_payload_fingerprint", json_fingerprint(payload))
    wb.save(output)
    emit(
        {
            "state": final_state,
            "output": str(output),
            "candidate_fingerprint": current_candidate_fp,
            "unique_clinical_rows": len(unique_sequences),
            "decisions_imported": len(assigned),
            "pending_unique_clinical_rows": len(pending),
            "pending_sequences": sorted(pending),
        }
    )


def unique_field_values(ws, rows: list[int], col: int) -> list[str]:
    values = []
    seen = set()
    for row in rows:
        value = clean_text(ws.cell(row, col).value, missing="")
        if value not in seen:
            seen.add(value)
            values.append(value)
    return values


def finalization_preview(config: dict[str, Any], workbook: str | Path) -> tuple[dict[str, Any], dict[str, Any]]:
    wb = load_workbook(workbook, data_only=False)
    draft_name = config["output"]["draft_sheet"]
    relation_name = config["output"]["relation_sheet"]
    if draft_name not in wb.sheetnames or relation_name not in wb.sheetnames or META_SHEET not in wb.sheetnames:
        raise ValueError("Workbook lacks the draft, relation, or metadata sheet.")
    draft = wb[draft_name]
    relation = wb[relation_name]
    draft_headers, _, draft_records = records_from_sheet(draft)
    expected_headers = [item["name"] for item in wide_schema(config)]
    if draft_headers != expected_headers:
        raise ValueError("Draft headers do not equal the approved Run Schema.")
    meta_map, meta_state = read_meta(wb[META_SHEET])
    require_matching_run(config, meta_state)
    if meta_state.get("draft_fingerprint") != sheet_fingerprint(draft):
        raise ValueError("Draft fingerprint differs from the workbook manifest.")
    current_candidate_fp = candidate_fingerprint(relation)
    if meta_state.get("candidate_fingerprint") != current_candidate_fp:
        raise ValueError("Candidate fingerprint differs from the workbook manifest.")
    relation_map = relation_header_map(relation)
    relation_last_row, _ = used_bounds(relation)
    groups: dict[str, list[int]] = defaultdict(list)
    for row in range(2, relation_last_row + 1):
        groups[stable_key(relation.cell(row, relation_map["临床分子序号"]).value)].append(row)
    fields = wide_fields(config)
    sequence_field = fields["sequence"]
    listed_source = config["output"]["listed_source_value"]
    clinical_source = config["output"]["clinical_source_value"]
    listed_records = [record for record in draft_records if record_source(record, fields, meta_map) == listed_source]
    clinical_records = [record for record in draft_records if record_source(record, fields, meta_map) == clinical_source]
    listed_sequences = {stable_key(record[sequence_field]) for record in listed_records}
    clinical_sequences = {stable_key(record[sequence_field]) for record in clinical_records}
    errors: list[str] = []
    missing_groups = sorted(clinical_sequences - set(groups))
    extra_groups = sorted(set(groups) - clinical_sequences)
    if missing_groups:
        errors.append(f"Clinical rows missing relation groups: {missing_groups}")
    if extra_groups:
        errors.append(f"Relation groups lack clinical draft rows: {extra_groups}")
    decisions: dict[str, dict[str, str]] = {}
    action_counts: Counter[str] = Counter()
    open_sequences: list[str] = []
    for sequence, rows in groups.items():
        values = {}
        for field in ("最终动作代码", "最终承接上市分子序号", "判断状态", "Disease修正值", "资产名称修正值"):
            unique = unique_field_values(relation, rows, relation_map[field])
            if len(unique) > 1:
                errors.append(f"Conflicting {field} values for clinical sequence {sequence}: {unique}")
            values[field] = unique[0] if unique else ""
        action = values["最终动作代码"]
        status = values["判断状态"]
        carrier = values["最终承接上市分子序号"]
        if action not in ALLOWED_ACTIONS:
            errors.append(f"Missing or unsupported action for clinical sequence {sequence}: {action}")
        if action == "HOLD" or not status or status.startswith("待"):
            open_sequences.append(sequence)
        if action == "DELETE_LISTED_COVERED":
            carriers = parse_carrier_sequences(carrier)
            invalid_carriers = [item for item in carriers if item not in listed_sequences]
            if not carriers or invalid_carriers:
                errors.append(f"Invalid listed carrier(s) for clinical sequence {sequence}: {carrier}")
        if action and action != "HOLD":
            action_counts[action] += 1
        decisions[sequence] = {
            "action": action,
            "carrier": carrier,
            "status": status,
            "disease_correction": values["Disease修正值"],
            "asset_correction": values["资产名称修正值"],
        }
    if open_sequences:
        errors.append(f"Open/HOLD clinical sequences remain: {sorted(open_sequences)}")
    keep = action_counts["KEEP"]
    expected_final = len(listed_records) + keep
    preview = {
        "state": "DECISIONS_CLOSED" if not errors else "DECISIONS_PENDING",
        "listed_rows": len(listed_records),
        "clinical_rows": len(clinical_records),
        "unique_relation_groups": len(groups),
        "candidate_rows": relation_last_row - 1,
        "action_counts": dict(action_counts),
        "expected_final_rows": expected_final,
        "expected_final_source_counts": {listed_source: len(listed_records), clinical_source: keep},
        "run_fingerprint": meta_state.get("run_fingerprint"),
        "candidate_fingerprint": current_candidate_fp,
        "pending_or_errors": errors,
    }
    context = {
        "wb": wb,
        "draft": draft,
        "relation": relation,
        "draft_headers": draft_headers,
        "draft_records": draft_records,
        "relation_map": relation_map,
        "groups": groups,
        "decisions": decisions,
        "listed_sequences": listed_sequences,
        "listed_records": listed_records,
        "clinical_records": clinical_records,
        "meta_map": meta_map,
    }
    return preview, context


def copy_row_style(source_ws, source_row: int, target_ws, target_row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        source = source_ws.cell(source_row, col)
        target = target_ws.cell(target_row, col)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def count_formulas(ws) -> int:
    last_row, last_col = used_bounds(ws)
    return sum(
        1
        for row in range(1, last_row + 1)
        for col in range(1, last_col + 1)
        if isinstance(ws.cell(row, col).value, str) and ws.cell(row, col).value.startswith("=")
    )


def command_finalize_full(args) -> None:
    config = load_json(args.config)
    validate_config_shape(config)
    if not config.get("schema_approved"):
        raise ValueError("Finalization requires an approved Run Schema.")
    workbook = Path(args.workbook)
    if not workbook.is_file():
        raise ValueError(f"Decided workbook was not found: {workbook}")
    preview, context = finalization_preview(config, workbook)
    if args.mode == "preview":
        emit(preview)
        return
    if not args.confirmed:
        raise ValueError("Generate mode requires --confirmed after preview and human approval.")
    if preview["pending_or_errors"]:
        raise ValueError("Finalization preview failed: " + " | ".join(preview["pending_or_errors"]))
    if not args.output:
        raise ValueError("Generate mode requires --output.")
    output = require_new_output(args.output, [workbook])
    context["wb"].close()
    shutil.copy2(workbook, output)
    wb = load_workbook(output, data_only=False)
    draft_name = config["output"]["draft_sheet"]
    relation_name = config["output"]["relation_sheet"]
    final_name = config["output"]["full_sheet"]
    if final_name in wb.sheetnames:
        output.unlink(missing_ok=True)
        raise ValueError(f"Final sheet already exists: {final_name}")
    draft = wb[draft_name]
    relation = wb[relation_name]
    _, _, draft_records = records_from_sheet(draft)
    relation_map = relation_header_map(relation)
    relation_last_row, _ = used_bounds(relation)
    groups: dict[str, list[int]] = defaultdict(list)
    for row in range(2, relation_last_row + 1):
        groups[stable_key(relation.cell(row, relation_map["临床分子序号"]).value)].append(row)
    decisions: dict[str, dict[str, str]] = {}
    for sequence, rows in groups.items():
        decisions[sequence] = {
            "action": unique_field_values(relation, rows, relation_map["最终动作代码"])[0],
            "carrier": (unique_field_values(relation, rows, relation_map["最终承接上市分子序号"]) or [""])[0],
            "disease_correction": (unique_field_values(relation, rows, relation_map["Disease修正值"]) or [""])[0],
            "asset_correction": (unique_field_values(relation, rows, relation_map["资产名称修正值"]) or [""])[0],
        }
    fields = wide_fields(config)
    sequence_field = fields["sequence"]
    listed_source = config["output"]["listed_source_value"]
    clinical_source = config["output"]["clinical_source_value"]
    meta_map, _ = read_meta(wb[META_SHEET])
    headers = [item["name"] for item in wide_schema(config)]
    final_rows: list[tuple[dict[str, Any], int, str, str]] = []
    clinical_to_final: dict[str, int] = {}
    for record in draft_records:
        source_value = record_source(record, fields, meta_map)
        old_sequence = stable_key(record[sequence_field])
        if source_value == listed_source:
            final_rows.append((record, record["__row_number__"], old_sequence, source_value))
        elif source_value == clinical_source and decisions[old_sequence]["action"] == "KEEP":
            corrected = dict(record)
            if decisions[old_sequence]["disease_correction"]:
                corrected[fields["disease"]] = decisions[old_sequence]["disease_correction"]
            if decisions[old_sequence]["asset_correction"]:
                corrected[fields["asset"]] = decisions[old_sequence]["asset_correction"]
            final_rows.append((corrected, record["__row_number__"], old_sequence, source_value))
    final = wb.create_sheet(final_name, 0)
    for col, header in enumerate(headers, start=1):
        final.cell(1, col).value = header
        source_cell = draft.cell(1, col)
        if source_cell.has_style:
            final.cell(1, col)._style = copy(source_cell._style)
        final.column_dimensions[get_column_letter(col)].width = draft.column_dimensions[get_column_letter(col)].width
    for new_sequence, (record, source_row, old_sequence, source_value) in enumerate(final_rows, start=1):
        record[sequence_field] = new_sequence
        output_row = new_sequence + 1
        for col, header in enumerate(headers, start=1):
            final.cell(output_row, col).value = record.get(header)
        copy_row_style(draft, source_row, final, output_row, len(headers))
        if source_value == clinical_source:
            clinical_to_final[old_sequence] = new_sequence
    final.freeze_panes = "A2"
    final.auto_filter.ref = final.dimensions
    for row in range(2, relation_last_row + 1):
        sequence = stable_key(relation.cell(row, relation_map["临床分子序号"]).value)
        action = decisions[sequence]["action"]
        if action == "KEEP":
            result = "已保留"
            final_sequence = clinical_to_final[sequence]
            sync = "已同步并保留" if decisions[sequence]["disease_correction"] or decisions[sequence]["asset_correction"] else "阶段三已执行：保留"
        elif action == "DELETE_LISTED_COVERED":
            result = "已删除：上市候选承接"
            final_sequence = ""
            sync = "阶段三已执行：删除临床行"
        else:
            result = "已排除"
            final_sequence = ""
            sync = "阶段三已执行：排除"
        relation.cell(row, relation_map["阶段执行结果"]).value = result
        relation.cell(row, relation_map["最终分子序号"]).value = final_sequence
        relation.cell(row, relation_map["主表同步状态"]).value = sync

    _, meta_state = read_meta(wb[META_SHEET])
    expected_listed_fp = meta_state.get("listed_raw_fingerprint", "")
    expected_clinical_fp = meta_state.get("clinical_raw_fingerprint", "")
    wb.remove(wb[draft_name])
    full_fp = sheet_fingerprint(final)
    supplement_records = [record for record, _, _, _ in final_rows if supplement_condition_matches(config, record)]
    supplement_enabled = bool(config.get("supplement", {}).get("enabled"))
    supplement_state = "SUPPLEMENT_PENDING" if supplement_enabled and supplement_records else "SUPPLEMENT_NOT_APPLICABLE"
    set_meta_value(wb[META_SHEET], "state", "FULL_BASELINE_LOCKED")
    set_meta_value(wb[META_SHEET], "full_baseline_fingerprint", full_fp)
    set_meta_value(wb[META_SHEET], "full_fingerprint", full_fp)
    set_meta_value(wb[META_SHEET], "clean_state", "NOT_GENERATED")
    set_meta_value(wb[META_SHEET], "clean_bound_full_fingerprint", "")
    set_meta_value(wb[META_SHEET], "supplement_state", supplement_state)
    set_meta_value(wb[META_SHEET], "supplement_candidate_count", len(supplement_records))
    set_meta_value(
        wb[META_SHEET],
        "module_states",
        json.dumps({item["name"]: item.get("status", "pending") for item in config.get("downstream_modules", [])}, ensure_ascii=False),
    )
    wb[META_SHEET].sheet_state = "hidden"
    desired = [
        final_name,
        relation_name,
        config["listed"]["raw_output_sheet"],
        config["clinical"]["raw_output_sheet"],
        META_SHEET,
    ]
    if any(name not in wb.sheetnames for name in desired):
        output.unlink(missing_ok=True)
        raise ValueError("Final workbook lacks a required output sheet.")
    extras = [name for name in wb.sheetnames if name not in desired]
    if extras:
        output.unlink(missing_ok=True)
        raise ValueError(f"Unexpected sheets would violate the full-baseline output contract: {extras}")
    wb._sheets = [wb[name] for name in desired]
    wb.save(output)

    verify = load_workbook(output, data_only=False)
    final_ws = verify[final_name]
    final_headers, _, final_records = records_from_sheet(final_ws)
    final_sequences = [record[sequence_field] for record in final_records]
    source_counts = Counter(source_value for _, _, _, source_value in final_rows)
    listed_fp = sheet_fingerprint(verify[config["listed"]["raw_output_sheet"]], include_title=False)
    clinical_fp = sheet_fingerprint(verify[config["clinical"]["raw_output_sheet"]], include_title=False)
    checks = {
        "sheet_order": verify.sheetnames == desired,
        "wide_schema": final_headers == headers,
        "final_rows": len(final_records) == preview["expected_final_rows"],
        "sequence": final_sequences == list(range(1, len(final_records) + 1)),
        "source_counts": dict(source_counts) == preview["expected_final_source_counts"],
        "listed_raw_fingerprint": listed_fp == expected_listed_fp,
        "clinical_raw_fingerprint": clinical_fp == expected_clinical_fp,
        "final_formula_cells": count_formulas(final_ws) == 0,
        "no_trailing_blank_rows": final_ws.max_row == len(final_records) + 1,
        "full_fingerprint": sheet_fingerprint(final_ws) == full_fp,
    }
    if not all(checks.values()):
        output.unlink(missing_ok=True)
        raise ValueError("Final workbook QC failed: " + json.dumps(checks, ensure_ascii=False))
    emit(
        {
            "state": "FULL_QC_PASSED",
            "output": str(output),
            "listed_rows": preview["listed_rows"],
            "clinical_rows": preview["clinical_rows"],
            "action_counts": preview["action_counts"],
            "final_rows": len(final_records),
            "final_source_counts": dict(source_counts),
            "final_sequence": f"1-{len(final_records)}",
            "pending": 0,
            "supplement_state": supplement_state,
            "supplement_candidate_count": len(supplement_records),
            "full_fingerprint": full_fp,
            "checks": checks,
        }
    )


def condition_matches(condition: dict[str, Any], record: dict[str, Any], missing: str) -> bool:
    if "all" in condition:
        items = condition["all"]
        if not isinstance(items, list):
            raise ValueError("Condition 'all' must be a list.")
        return all(condition_matches(item, record, missing) for item in items)
    if "any" in condition:
        items = condition["any"]
        if not isinstance(items, list):
            raise ValueError("Condition 'any' must be a list.")
        return any(condition_matches(item, record, missing) for item in items)
    if "not" in condition:
        return not condition_matches(condition["not"], record, missing)
    field = clean_text(condition.get("field"), missing="")
    if not field:
        raise ValueError("Leaf condition lacks a field.")
    operation = condition.get("op", "eq")
    raw = record.get(field)
    actual = clean_text(raw, missing)
    if operation == "has_value":
        return bool(actual)
    if operation == "is_blank":
        return not actual
    if operation == "eq":
        return actual == clean_text(condition.get("value"), missing)
    if operation == "ne":
        return actual != clean_text(condition.get("value"), missing)
    if operation in {"in", "not_in"}:
        values = condition.get("values")
        if not isinstance(values, list):
            raise ValueError(f"Condition {operation} requires a values list.")
        normalized = {clean_text(value, missing) for value in values}
        result = actual in normalized
        return result if operation == "in" else not result
    if operation == "contains":
        return clean_text(condition.get("value"), missing) in actual
    raise ValueError(f"Unsupported condition operation: {operation}")


def supplement_condition_matches(config: dict[str, Any], record: dict[str, Any]) -> bool:
    supplement = config.get("supplement", {})
    if not supplement.get("enabled"):
        return False
    return condition_matches(supplement["trigger"], record, config.get("missing_display", "—"))


def clean_rule_value(rule: dict[str, Any], record: dict[str, Any], missing: str) -> Any:
    mode = rule["mode"]
    if mode == "constant":
        return display_value(rule.get("value"), missing)
    if mode == "missing":
        return missing
    if mode in {"column", "column_or_missing"}:
        column = rule["column"]
        if column not in record:
            if mode == "column_or_missing":
                return missing
            raise ValueError(f"Required full-pool field is missing during Clean derivation: {column}")
        return display_value(record.get(column), missing)
    if mode == "coalesce_columns":
        for column in rule["columns"]:
            if column in record and not is_blank(record.get(column)) and clean_text(record.get(column), missing):
                return record.get(column)
        return missing
    if mode == "conditional":
        for case in rule["cases"]:
            if condition_matches(case["when"], record, missing):
                return clean_rule_value(case["rule"], record, missing)
        return clean_rule_value(rule["default"], record, missing)
    raise ValueError(f"Unsupported Clean mapping mode: {mode}")


def load_full_context(config: dict[str, Any], workbook: str | Path):
    path = Path(workbook)
    if not path.is_file():
        raise ValueError(f"Full-pool workbook was not found: {path}")
    wb = load_workbook(path, data_only=False)
    full_name = config["output"]["full_sheet"]
    if full_name not in wb.sheetnames or META_SHEET not in wb.sheetnames:
        raise ValueError("Workbook lacks the full-pool or workflow metadata sheet.")
    full = wb[full_name]
    headers, mapping, records = records_from_sheet(full)
    expected_headers = [item["name"] for item in wide_schema(config)]
    if headers != expected_headers:
        raise ValueError("Full-pool headers differ from the approved wide_schema.")
    meta = meta_values(wb[META_SHEET])
    if meta.get("config_fingerprint") != config_fingerprint(config):
        raise ValueError("Run config contract fingerprint differs from the full-pool manifest.")
    sequence_field = wide_fields(config)["sequence"]
    sequences = [record.get(sequence_field) for record in records]
    if sequences != list(range(1, len(records) + 1)):
        raise ValueError("Full-pool sequence is not the locked 1..K contract.")
    if count_formulas(full):
        raise ValueError("Full pool contains formulas; static values are required before supplement/Clean work.")
    return wb, full, headers, mapping, records, meta, sheet_fingerprint(full)


def module_state_map(config: dict[str, Any]) -> dict[str, str]:
    return {item["name"]: item.get("status", "pending") for item in config.get("downstream_modules", [])}


def command_build_supplement_template(args) -> None:
    config = load_json(args.config)
    validate_config_shape(config)
    supplement = config.get("supplement", {})
    if not supplement.get("enabled"):
        raise ValueError("Supplementation is disabled in the run config.")
    wb, _, _, _, records, _, full_fp = load_full_context(config, args.workbook)
    output = require_new_output(args.output, [args.workbook])
    candidates = [record for record in records if supplement_condition_matches(config, record)]
    id_field = supplement["id_field"]
    ids = [stable_key(record.get(id_field)) for record in candidates]
    if any(not value for value in ids) or len(ids) != len(set(ids)):
        raise ValueError("Supplement candidates contain blank or duplicate locked full IDs.")
    context_fields = supplement.get("context_fields", [])
    target_fields = supplement.get("target_fields", [])
    status_field = supplement["status_field"]
    note_field = supplement.get("note_field", "补充说明")
    headers = []
    for field in [id_field] + context_fields + target_fields + [status_field, note_field]:
        if field not in headers:
            headers.append(field)
    out_wb = Workbook()
    ws = out_wb.active
    ws.title = supplement["template_sheet"]
    ws.append(headers)
    missing = config.get("missing_display", "—")
    for record in candidates:
        row = []
        for header in headers:
            if header == status_field:
                row.append(supplement["open_status"])
            elif header == note_field:
                row.append("")
            else:
                value = record.get(header)
                row.append("" if not clean_text(value, missing) else value)
        ws.append(row)
    style_table(ws, header_fill="548235")
    meta_ws = out_wb.create_sheet("__supplement_meta")
    meta_ws.append(["config_fingerprint", config_fingerprint(config)])
    meta_ws.append(["source_full_fingerprint", full_fp])
    meta_ws.append(["candidate_id_fingerprint", json_fingerprint(ids)])
    meta_ws.append(["candidate_count", len(ids)])
    meta_ws.sheet_state = "hidden"
    out_wb.save(output)
    verify = load_workbook(output, data_only=False)
    _, _, verify_records = records_from_sheet(verify[supplement["template_sheet"]])
    if len(verify_records) != len(candidates):
        output.unlink(missing_ok=True)
        raise ValueError("Supplement-template row verification failed.")
    emit(
        {
            "state": "SUPPLEMENT_TEMPLATE_BUILT",
            "output": str(output),
            "candidate_count": len(candidates),
            "locked_ids": ids,
            "source_full_fingerprint": full_fp,
            "target_fields": target_fields,
        }
    )
    wb.close()


def existing_supplement_audit(wb, sheet_name: str, id_field: str) -> dict[str, dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return {}
    _, _, records = records_from_sheet(wb[sheet_name])
    result: dict[str, dict[str, Any]] = {}
    for record in records:
        key = stable_key(record.get(id_field))
        if key:
            result[key] = record
    return result


def command_apply_supplements(args) -> None:
    config = load_json(args.config)
    validate_config_shape(config)
    supplement = config.get("supplement", {})
    if not supplement.get("enabled"):
        raise ValueError("Supplementation is disabled in the run config.")
    source = Path(args.workbook)
    supplement_path = Path(args.supplement)
    if not supplement_path.is_file():
        raise ValueError(f"Supplement workbook was not found: {supplement_path}")
    output = require_new_output(args.output, [source, supplement_path])
    shutil.copy2(source, output)
    wb, full, _, full_mapping, full_records, meta, before_fp = load_full_context(config, output)
    supp_wb = load_workbook(supplement_path, data_only=True)
    template_sheet = supplement["template_sheet"]
    if template_sheet not in supp_wb.sheetnames:
        output.unlink(missing_ok=True)
        raise ValueError(f"Supplement sheet was not found: {template_sheet}")
    supp_headers, _, returned = records_from_sheet(supp_wb[template_sheet])
    id_field = supplement["id_field"]
    status_field = supplement["status_field"]
    note_field = supplement.get("note_field", "补充说明")
    target_fields = supplement.get("target_fields", [])
    required_headers = [id_field, status_field] + target_fields
    missing_headers = [field for field in required_headers if field not in supp_headers]
    if missing_headers:
        output.unlink(missing_ok=True)
        raise ValueError("Supplement workbook lacks required fields: " + ", ".join(missing_headers))
    returned_ids = [stable_key(record.get(id_field)) for record in returned]
    if any(not value for value in returned_ids) or len(returned_ids) != len(set(returned_ids)):
        output.unlink(missing_ok=True)
        raise ValueError("Returned supplement IDs are blank or duplicated.")
    candidates = [record for record in full_records if supplement_condition_matches(config, record)]
    candidate_by_id = {stable_key(record.get(id_field)): record for record in candidates}
    full_by_id = {stable_key(record.get(id_field)): record for record in full_records}
    unknown = sorted(set(returned_ids) - set(candidate_by_id))
    if unknown:
        output.unlink(missing_ok=True)
        raise ValueError(f"Supplement rows contain unknown/noncandidate locked IDs: {unknown}")
    audit_sheet = supplement.get("audit_sheet", DEFAULT_SUPPLEMENT_AUDIT_SHEET)
    prior_audit = existing_supplement_audit(wb, audit_sheet, id_field)
    returned_by_id = {stable_key(record.get(id_field)): record for record in returned}
    allowed_statuses = {supplement["open_status"], *supplement["closed_statuses"]}
    missing = config.get("missing_display", "—")
    changed_fields: set[str] = set()
    changed_ids: set[str] = set()
    final_statuses: dict[str, str] = {}
    notes: dict[str, Any] = {}
    for candidate_id, record in candidate_by_id.items():
        returned_record = returned_by_id.get(candidate_id)
        prior = prior_audit.get(candidate_id, {})
        status = clean_text(
            returned_record.get(status_field) if returned_record else prior.get(status_field),
            missing,
        ) or supplement["open_status"]
        if status not in allowed_statuses:
            output.unlink(missing_ok=True)
            raise ValueError(f"Unsupported supplement status for ID {candidate_id}: {status}")
        final_statuses[candidate_id] = status
        notes[candidate_id] = (returned_record or prior).get(note_field, "")
        if returned_record:
            row_number = record["__row_number__"]
            for field in target_fields:
                incoming = returned_record.get(field)
                if not clean_text(incoming, missing):
                    continue
                existing = full.cell(row_number, full_mapping[field]).value
                if clean_text(existing, missing) and clean_text(existing, missing) != clean_text(incoming, missing) and not args.allow_overwrite:
                    output.unlink(missing_ok=True)
                    raise ValueError(f"Supplement would overwrite a valid value for ID {candidate_id}, field {field}.")
                if clean_text(existing, missing) != clean_text(incoming, missing):
                    full.cell(row_number, full_mapping[field]).value = incoming
                    record[field] = incoming
                    changed_fields.add(field)
                    changed_ids.add(candidate_id)
        if status == "已补充" and not any(clean_text(record.get(field), missing) for field in target_fields):
            output.unlink(missing_ok=True)
            raise ValueError(f"ID {candidate_id} is marked 已补充 but has no usable target-field value.")

    if audit_sheet in wb.sheetnames:
        wb.remove(wb[audit_sheet])
    audit = wb.create_sheet(audit_sheet, len(wb.sheetnames) - 1)
    audit_headers = []
    for field in [id_field] + supplement.get("context_fields", []) + target_fields + [status_field, note_field]:
        if field not in audit_headers:
            audit_headers.append(field)
    audit.append(audit_headers)
    for record in candidates:
        candidate_id = stable_key(record.get(id_field))
        audit.append(
            [
                final_statuses[candidate_id] if field == status_field else notes[candidate_id] if field == note_field else record.get(field)
                for field in audit_headers
            ]
        )
    audit.sheet_state = "hidden"
    closed_statuses = set(supplement["closed_statuses"])
    pending_ids = sorted(candidate_id for candidate_id, status in final_statuses.items() if status not in closed_statuses)
    supplement_state = "SUPPLEMENT_CLOSED" if not pending_ids else "SUPPLEMENT_PENDING"
    after_fp = sheet_fingerprint(full)
    affected_modules = []
    for module in config.get("downstream_modules", []):
        if changed_fields & set(module.get("affected_by", [])):
            affected_modules.append(module["name"])
    set_meta_value(wb[META_SHEET], "full_fingerprint", after_fp)
    set_meta_value(wb[META_SHEET], "supplement_state", supplement_state)
    set_meta_value(wb[META_SHEET], "supplement_candidate_count", len(candidates))
    set_meta_value(wb[META_SHEET], "supplement_pending_count", len(pending_ids))
    set_meta_value(wb[META_SHEET], "supplement_input_sha256", file_sha256(supplement_path))
    if changed_fields and config["output"]["clean_sheet"] in wb.sheetnames:
        set_meta_value(wb[META_SHEET], "clean_state", "CLEAN_STALE")
    set_meta_value(wb[META_SHEET], "affected_modules_after_supplement", json.dumps(affected_modules, ensure_ascii=False))
    wb[META_SHEET].sheet_state = "hidden"
    wb.save(output)
    verify = load_workbook(output, data_only=False)
    if sheet_fingerprint(verify[config["output"]["full_sheet"]]) != after_fp:
        output.unlink(missing_ok=True)
        raise ValueError("Full-pool fingerprint verification failed after supplement application.")
    emit(
        {
            "state": supplement_state,
            "output": str(output),
            "candidate_count": len(candidates),
            "returned_count": len(returned),
            "pending_count": len(pending_ids),
            "pending_ids": pending_ids,
            "changed_ids": sorted(changed_ids),
            "changed_fields": sorted(changed_fields),
            "affected_modules": affected_modules,
            "full_fingerprint_before": before_fp,
            "full_fingerprint_after": after_fp,
            "clean_state": meta_values(verify[META_SHEET]).get("clean_state"),
        }
    )


def command_derive_clean(args) -> None:
    config = load_json(args.config)
    validate_config_shape(config)
    source = Path(args.workbook)
    output = require_new_output(args.output, [source])
    shutil.copy2(source, output)
    wb, full, _, _, records, meta, full_fp = load_full_context(config, output)
    stage = args.stage
    supplement = config.get("supplement", {})
    module_states = module_state_map(config)
    blockers: list[str] = []
    if stage == "final":
        if not args.confirmed_current_full:
            blockers.append("Final Clean requires --confirmed-current-full.")
        if supplement.get("enabled") and supplement.get("required_for_final_clean"):
            if meta.get("supplement_state") not in {"SUPPLEMENT_CLOSED", "SUPPLEMENT_NOT_APPLICABLE"}:
                blockers.append(f"Required supplement state is open: {meta.get('supplement_state') or 'UNKNOWN'}")
        closed_module_states = {"completed", "not_applicable", "approved_skip"}
        for module in config.get("downstream_modules", []):
            if module.get("required_for_final_clean") and module.get("status", "pending") not in closed_module_states:
                blockers.append(f"Required downstream module is open: {module['name']}")
    if blockers:
        output.unlink(missing_ok=True)
        raise ValueError("Clean derivation gate failed: " + " | ".join(blockers))

    include_auxiliary = bool(args.include_auxiliary or config.get("include_clean_auxiliary_by_default", False))
    if stage == "final" and not config.get("include_clean_auxiliary_in_final", False):
        include_auxiliary = False
    schema = list(config["clean_schema"])
    if include_auxiliary:
        schema.extend(config.get("clean_auxiliary_schema", []))
    clean_name = config["output"]["clean_sheet"]
    if clean_name in wb.sheetnames:
        wb.remove(wb[clean_name])
    full_index = wb.sheetnames.index(config["output"]["full_sheet"])
    clean_ws = wb.create_sheet(clean_name, full_index + 1)
    clean_headers = [item["name"] for item in schema]
    clean_ws.append(clean_headers)
    missing = config.get("missing_display", "—")
    for record in records:
        clean_ws.append([clean_rule_value({key: value for key, value in item.items() if key != "name"}, record, missing) for item in schema])
    style_table(clean_ws, header_fill="2F75B5")
    sequence_field = wide_fields(config)["sequence"]
    if sequence_field not in clean_headers:
        output.unlink(missing_ok=True)
        raise ValueError(f"Clean schema must retain the locked sequence field: {sequence_field}")
    _, _, clean_records = records_from_sheet(clean_ws)
    full_ids = [record[sequence_field] for record in records]
    clean_ids = [record[sequence_field] for record in clean_records]
    if full_ids != clean_ids:
        output.unlink(missing_ok=True)
        raise ValueError("Full and Clean ID/order contract failed before save.")
    clean_state = "CLEAN_FINAL" if stage == "final" else "CLEAN_WORKING"
    set_meta_value(wb[META_SHEET], "state", clean_state)
    set_meta_value(wb[META_SHEET], "full_fingerprint", full_fp)
    set_meta_value(wb[META_SHEET], "clean_state", clean_state)
    set_meta_value(wb[META_SHEET], "clean_bound_full_fingerprint", full_fp)
    set_meta_value(wb[META_SHEET], "clean_schema_used", json.dumps(clean_headers, ensure_ascii=False))
    set_meta_value(wb[META_SHEET], "module_states", json.dumps(module_states, ensure_ascii=False))
    wb[META_SHEET].sheet_state = "hidden"
    wb.save(output)
    verify = load_workbook(output, data_only=False)
    verify_full = verify[config["output"]["full_sheet"]]
    verify_clean = verify[clean_name]
    verify_clean_headers, _, verify_clean_records = records_from_sheet(verify_clean)
    verify_clean_ids = [record[sequence_field] for record in verify_clean_records]
    checks = {
        "full_fingerprint": sheet_fingerprint(verify_full) == full_fp,
        "clean_headers": verify_clean_headers == clean_headers,
        "row_count": len(verify_clean_records) == len(records),
        "id_order": verify_clean_ids == full_ids,
        "clean_formula_cells": count_formulas(verify_clean) == 0,
        "no_trailing_blank_rows": verify_clean.max_row == len(records) + 1,
        "clean_bound_to_current_full": meta_values(verify[META_SHEET]).get("clean_bound_full_fingerprint") == full_fp,
    }
    if not all(checks.values()):
        output.unlink(missing_ok=True)
        raise ValueError("Clean workbook QC failed: " + json.dumps(checks, ensure_ascii=False))
    emit(
        {
            "state": clean_state,
            "output": str(output),
            "full_rows": len(records),
            "clean_rows": len(verify_clean_records),
            "clean_fields": len(clean_headers),
            "auxiliary_fields_included": include_auxiliary,
            "supplement_state": meta.get("supplement_state"),
            "module_states": module_states,
            "full_fingerprint": full_fp,
            "checks": checks,
        }
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Merge listed and clinical molecule pools through gated workbook stages.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect", help="Read-only workbook and schema preflight.")
    inspect_parser.add_argument("--config", required=True)
    inspect_parser.set_defaults(func=command_inspect)

    draft_parser = subparsers.add_parser("build-draft", help="Build an approved-schema combined draft.")
    draft_parser.add_argument("--config", required=True)
    draft_parser.add_argument("--output", required=True)
    draft_parser.add_argument("--confirmed-schema", action="store_true")
    draft_parser.set_defaults(func=command_build_draft)

    candidate_parser = subparsers.add_parser("build-candidates", help="Generate high-recall cross-library candidates.")
    candidate_parser.add_argument("--config", required=True)
    candidate_parser.add_argument("--workbook", required=True)
    candidate_parser.add_argument("--output", required=True)
    candidate_parser.set_defaults(func=command_build_candidates)

    decision_parser = subparsers.add_parser("import-decisions", help="Import structured decisions into every candidate row in a clinical group.")
    decision_parser.add_argument("--config", required=True)
    decision_parser.add_argument("--workbook", required=True)
    decision_parser.add_argument("--decisions", required=True)
    decision_parser.add_argument("--output", required=True)
    decision_parser.set_defaults(func=command_import_decisions)

    final_parser = subparsers.add_parser("finalize-full", help="Preview or generate the authoritative full-pool baseline.")
    final_parser.add_argument("--mode", choices=("preview", "generate"), required=True)
    final_parser.add_argument("--config", required=True)
    final_parser.add_argument("--workbook", required=True)
    final_parser.add_argument("--output")
    final_parser.add_argument("--confirmed", action="store_true")
    final_parser.set_defaults(func=command_finalize_full)

    supplement_template_parser = subparsers.add_parser(
        "build-supplement-template",
        help="Export clinical-source marketed rows for optional listed-information supplementation.",
    )
    supplement_template_parser.add_argument("--config", required=True)
    supplement_template_parser.add_argument("--workbook", required=True)
    supplement_template_parser.add_argument("--output", required=True)
    supplement_template_parser.set_defaults(func=command_build_supplement_template)

    supplement_apply_parser = subparsers.add_parser(
        "apply-supplements",
        help="Apply returned marketed information to the full pool by locked full ID.",
    )
    supplement_apply_parser.add_argument("--config", required=True)
    supplement_apply_parser.add_argument("--workbook", required=True)
    supplement_apply_parser.add_argument("--supplement", required=True)
    supplement_apply_parser.add_argument("--output", required=True)
    supplement_apply_parser.add_argument("--allow-overwrite", action="store_true")
    supplement_apply_parser.set_defaults(func=command_apply_supplements)

    clean_parser = subparsers.add_parser("derive-clean", help="Derive a Working or Final Clean view from the current full pool.")
    clean_parser.add_argument("--stage", choices=("working", "final"), required=True)
    clean_parser.add_argument("--config", required=True)
    clean_parser.add_argument("--workbook", required=True)
    clean_parser.add_argument("--output", required=True)
    clean_parser.add_argument("--include-auxiliary", action="store_true")
    clean_parser.add_argument("--confirmed-current-full", action="store_true")
    clean_parser.set_defaults(func=command_derive_clean)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        args.func(args)
        return 0
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(json.dumps({"state": "ERROR", "error": str(exc)}, ensure_ascii=False, indent=2), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
